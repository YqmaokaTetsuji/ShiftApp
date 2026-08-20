import streamlit as st
import pandas as pd
import os
import datetime
import calendar
import unicodedata
from filelock import FileLock
import requests
import io
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
st.set_page_config(page_title='シフト希望提出フォーム', layout='wide')
# ==========================================
# 設定値の定義
# ==========================================
JST = datetime.timezone(datetime.timedelta(hours=+9), 'JST')
now_jst = datetime.datetime.now(JST)
today = now_jst.date()

TARGET_YEAR = today.year + 1 if today.month == 12 else today.year
TARGET_MONTH = 1 if today.month == 12 else today.month + 1

# バックアップ用のファイル指定
CSV_REQUESTS = f'【プログラム用】{TARGET_MONTH}月シフト提出状況.csv'
EXCEL_REQUESTS = f'【店長確認用】{TARGET_MONTH}月シフト提出状況.xlsx'
LOCK_FILE = f'{TARGET_MONTH}月シフト提出状況.lock'

# 部門の指定
DEPARTMENTS = ['選択してください', '季節AV', '家電', '情報', '通信']

# パスワードの指定
ADMIN_PASSWORD = "st.secrets['admin_password']"

# GoogleスプレッドシートのScriptsたち
GAS_URL = "st.secrets['gas_url']"

hide_streamlit_style = '''
<style>
#MainMenu {visibility: hidden;}
header {visibility: hidden;}
footer {visibility: hidden;}
</style>
'''
st.markdown(hide_streamlit_style, unsafe_allow_html=True)
# ==========================================
# 関数の定義
# ==========================================
def init_session_state():
    '''状態の初期化'''
    if 'confirm_mode' not in st.session_state:
        st.session_state.confirm_mode = False
    if 'is_submitted' not in st.session_state:
        st.session_state.is_submitted = False
    if 'excel_warning' not in st.session_state:
        st.session_state.excel_warning = False
    if 'is_processing' not in st.session_state:
        st.session_state.is_processing = False

@st.dialog('シフトの一括入力')
def bulk_input_dialog(day_labels):
    st.write('一括で入力したい曜日や日付を選択してください。')
    
    weekdays = ['月', '火', '水', '木', '金', '土', '日']
    selected_dows = st.multiselect('1．曜日でまとめて選択', weekdays, placeholder='（例：月、水、金）')
    selected_dates = st.multiselect('2．特定の日付を追加で選択', day_labels, placeholder='（例：15日、20日）')
    
    st.divider()
    
    shift_type = st.radio('適用するシフトを選択', ['希望なし', '休', '早', '遅', '時間指定', '有給'], horizontal=True)
    specific_time = ''
    if shift_type == '時間指定':
        specific_time = st.text_input('希望時間を入力（例：11-19, 早-15）', key='bulk_time_input')
        
    if st.button('この内容で一括反映する', type='primary', use_container_width=True):
        # 選択された日付と曜日を合体させる
        target_labels = set(selected_dates)
        for label in day_labels:
            for dow in selected_dows:
                if f'（{dow}）' in label:
                    target_labels.add(label)
                    
        if not target_labels:
            st.error('対象の日付または曜日を1つ以上選択してください。')
        else:
            # メイン画面の入力欄（セッションステート）を直接書き換える
            for label in target_labels:
                st.session_state[f'radio_{label}'] = shift_type
                if shift_type == '時間指定':
                    st.session_state[f'time_{label}'] = specific_time
            # 画面を更新してダイアログを閉じる
            st.rerun()

# 来月の日にちとその曜日を取得する関数
def get_month_days():
    _, num_days = calendar.monthrange(TARGET_YEAR, TARGET_MONTH)
    weekdays_ja = ['月', '火', '水', '木', '金', '土', '日']
    
    day_labels = []
    for d in range(1, num_days + 1):
        dt = datetime.date(TARGET_YEAR, TARGET_MONTH, d)
        day_labels.append(f'{d}日（{weekdays_ja[dt.weekday()]}）')
        
    return num_days, day_labels

# スプシにデータを送信
def save_shift_data(emp_code, name, department, target_days, shift_requests, remarks):
    safe_emp_code = unicodedata.normalize('NFKC', emp_code).strip()
    safe_name = name.strip()
    
    # 送信用のデータを作成
    data = {
        '対象月': f'{TARGET_MONTH}月',
        '提出日時': datetime.datetime.now(JST).strftime('%Y-%m-%d %H:%M:%S'),
        '従業員コード': safe_emp_code,
        '名前': safe_name, 
        '部門': department,
        '希望出勤時間': target_days
    }
    data.update(shift_requests)

    data['備考'] = remarks.strip()
    
    # Googleスプレッドシートへデータを送信
    try:
        response = requests.post(GAS_URL, json=data, timeout=10)
        if response.status_code != 200 or response.json().get('status') != 'success':
            return 'gas_error'
    except Exception:
        return 'gas_error'
    
    # 念のためサーバーのローカル環境にもバックアップ（今は過去の遺産と化した）
    df_submit = pd.DataFrame([data])
    lock = FileLock(LOCK_FILE)
    with lock:
        if os.path.exists(CSV_REQUESTS):
            try:
                df_existing = pd.read_csv(CSV_REQUESTS, encoding='utf-8-sig')
                df_final = pd.concat([df_existing, df_submit], ignore_index=True)
            except Exception:
                df_final = df_submit
        else:
            df_final = df_submit
            
        try:
            df_final.to_csv(CSV_REQUESTS, index=False, encoding='utf-8-sig')
            df_final.to_excel(EXCEL_REQUESTS, index=False)
        except Exception:
            pass # クラウド環境ではエラーを無視して進める
            
    return 'success'

# プレビュー用のカレンダーを作成、表示
def generate_styled_calendar(day_labels, shift_requests):
    calendar.setfirstweekday(calendar.SUNDAY)
    cal_matrix = calendar.monthcalendar(TARGET_YEAR, TARGET_MONTH)
    cal_data = []
    
    for week in cal_matrix:
        week_data = []
        for d in week:
            if d == 0:
                week_data.append('')
            else:
                label = day_labels[d-1]
                req = shift_requests.get(label, '')
                display_text = '出' if req == '' else req
                week_data.append(f'{d}日: {display_text}')
        cal_data.append(week_data)
        
    df = pd.DataFrame(cal_data, columns=['日', '月', '火', '水', '木', '金', '土'])

    # 日によってセルの設定を変えるよ
    def style_cells(data):
        styles = pd.DataFrame('', index=data.index, columns=data.columns)
        for row in data.index:
            for col in data.columns:
                val = str(data.loc[row, col])
                css = []
                if col == '土': css.append('background-color: #E6F2FF')
                elif col == '日': css.append('background-color: #FFE6E6')
                if '休' in val or '有' in val:
                    css.append('color: #FF0000; font-weight: bold')
                styles.loc[row, col] = '; '.join(css)
        return styles

    return df.style.apply(style_cells, axis=None)

# 店長用のメニューを表示（ここ冗長すぎるから絶対にリファクタリンぐ最優先！！！！！！！！！！！！！！！！！！！！！！！！！！！！！）
def show_admin_panel():
    with st.popover('店長専用メニュー', use_container_width=True):
        admin_pass = st.text_input('店長用パスワードを入力', type='password')
        if admin_pass == ADMIN_PASSWORD:
            st.write('---')
            st.markdown('#### シフトデータのダウンロード')
            st.write('スプレッドシートの最新データを、見やすく色付けされたExcelファイルとして保存します。')
            
            if st.button('最新のExcelを作成する', use_container_width=True):
                with st.spinner('クラウドからデータを取得＆Excelを装飾中...'):
                    try:
                        response = requests.get(f'{GAS_URL}?type=shift&month={TARGET_MONTH}', timeout=15)
                        if response.status_code == 200:
                            raw_data = response.json()
                            if len(raw_data) > 1:
                                df_dl = pd.DataFrame(raw_data[1:], columns=raw_data[0])
                                df_dl.columns = df_dl.columns.str.strip()
                                
                                # 不要な列をExcelの表から除外する
                                drop_cols = [c for c in ['対象月', '提出日時'] if c in df_dl.columns]
                                if drop_cols:
                                    df_dl = df_dl.drop(columns=drop_cols)
                                
                                # 部門順 従業員コード順に並び替え
                                if '部門' in df_dl.columns and '従業員コード' in df_dl.columns:
                                    dept_order = {dept: i for i, dept in enumerate(DEPARTMENTS) if dept != '選択してください'}
                                    df_dl['_sort_key'] = df_dl['部門'].map(lambda x: dept_order.get(x, 99))
                                    df_dl['_code_num'] = pd.to_numeric(df_dl['従業員コード'], errors='coerce').fillna(999999)
                                    df_dl = df_dl.sort_values(['_sort_key', '_code_num']).drop(columns=['_sort_key', '_code_num'])

                                    # 部門が切り替わるタイミングで空白行を挿入
                                    new_rows = []
                                    current_dept = None
                                    for idx, row in df_dl.iterrows():
                                        if current_dept is not None and current_dept != row['部門']:
                                            # 部門が変わったら、すべての列が空っぽの行を1つ挟む
                                            new_rows.append({col: '' for col in df_dl.columns})
                                        new_rows.append(row.to_dict())
                                        current_dept = row['部門']
                                    df_dl = pd.DataFrame(new_rows)

                                # 空中でExcelを作成 デザイン装飾
                                output = io.BytesIO()
                                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                    sheet_name = f'{TARGET_MONTH}月シフト提出'
                                    df_dl.to_excel(writer, index=False, sheet_name=sheet_name)
                                    
                                    # ここから openpyxl を使った自動デザイン装飾
                                    ws = writer.sheets[sheet_name]
                                    
                                    #左4列（従業員コード〜希望出勤時間を固定
                                    ws.freeze_panes = 'E2'
                                    
                                    # 色・罫線・フォントの準備
                                    fill_header = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid') # グレー
                                    fill_sat = PatternFill(start_color='E6F2FF', end_color='E6F2FF', fill_type='solid')    # 薄い青
                                    fill_sun = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')    # 薄い赤
                                    
                                    font_header = Font(name='メイリオ', size=10, bold=True)
                                    font_normal = Font(name='メイリオ', size=10)
                                    font_off = Font(name='メイリオ', size=10, bold=True, color='FF0000') # 赤字・太字
                                    
                                    border_thin = Border(
                                        left=Side(style='thin', color='D9D9D9'),
                                        right=Side(style='thin', color='D9D9D9'),
                                        top=Side(style='thin', color='D9D9D9'),
                                        bottom=Side(style='thin', color='D9D9D9')
                                    )
                                    
                                    # 全セルのスタイリング＆色付け
                                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                                        # この行が空白かどうかのチェック
                                        is_blank_row = False
                                        if row[0].row > 1:
                                            val_code = str(row[0].value or '').strip()
                                            val_name = str(row[1].value or '').strip()
                                            # 従業員コードも名前も空っぽなら、さっき追加した空白行だと判定
                                            if val_code == '' and val_name == '':
                                                is_blank_row = True

                                        for cell in row:
                                            # 空白の場合はスルー
                                            if is_blank_row:
                                                continue

                                            cell.border = border_thin
                                            col_name = str(ws.cell(row=1, column=cell.column).value or '')
                                            
                                            if cell.row == 1:
                                                # 見出し行のデザイン
                                                cell.fill = fill_header
                                                cell.font = font_header
                                                cell.alignment = Alignment(horizontal='center', vertical='center')
                                            else:
                                                # データ行：フォントと中央揃えの基本設定
                                                cell.font = font_normal
                                                cell.alignment = Alignment(horizontal='center', vertical='center')
                                                
                                                # 土曜と日曜塗り分ける
                                                if '（土）' in col_name:
                                                    cell.fill = fill_sat
                                                elif '（日）' in col_name:
                                                    cell.fill = fill_sun
                                                    
                                                # 「休」「有」の文字が入っていたら赤字にする
                                                if '（' in col_name and cell.value and ('休' in str(cell.value) or '有' in str(cell.value)):
                                                    cell.font = font_off
                                    
                                    # 列幅を見やすく自動調整
                                    for col in ws.columns:
                                        max_len = max(len(str(cell.value or '')) for cell in col)
                                        col_letter = get_column_letter(col[0].column)
                                        ws.column_dimensions[col_letter].width = max(max_len * 2, 12)

                                excel_data = output.getvalue()
                                st.success('✅ Excel準備完了！')
                                st.download_button(
                                    label='Excelファイル（.xlsx）を保存',
                                    data=excel_data,
                                    file_name=EXCEL_REQUESTS,
                                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                    use_container_width=True,
                                    type='primary'
                                )
                            else:
                                st.info('まだ誰もシフトを提出していません。')
                        else:
                            st.error('データの取得に失敗しました。')
                    except Exception as e:
                        st.error(f'通信エラーが発生しました: {e}')

            st.write('---')
            st.markdown('#### 👤 提出状況一覧（未提出チェック）')

            # 現在の提出状況を確認する
            with st.spinner('名簿と提出状況を照合中...'):
                try:
                    res_member = requests.get(f'{GAS_URL}?type=member', timeout=15)
                    res_shift = requests.get(f'{GAS_URL}?type=shift&month={TARGET_MONTH}', timeout=15)
                    
                    if res_member.status_code == 200 and res_shift.status_code == 200:
                        raw_member = res_member.json()
                        raw_shift = res_shift.json()
                        
                        if len(raw_member) > 1:
                            df_members = pd.DataFrame(raw_member[1:], columns=raw_member[0])
                            df_members.columns = df_members.columns.str.strip()
                            
                            if set(['従業員コード', '名前', '部門']).issubset(df_members.columns):
                                df_status = df_members[['従業員コード', '名前', '部門']].copy()
                                df_status['従業員コード'] = df_status['従業員コード'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
                                
                                submitted_codes = []
                                if len(raw_shift) > 1:
                                    df_submitted = pd.DataFrame(raw_shift[1:], columns=raw_shift[0])
                                    df_submitted.columns = df_submitted.columns.str.strip()
                                    if '従業員コード' in df_submitted.columns:
                                        submitted_codes = df_submitted['従業員コード'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip().tolist()
                                
                                df_status['提出状況'] = df_status['従業員コード'].apply(
                                    lambda x: '提出済' if x in submitted_codes else '未提出'
                                )
                                df_status = df_status.sort_values('提出状況', ascending=False)
                                
                                def highlight_unsubmitted(row):
                                    return ['background-color: #FFE6E6' if row['提出状況'] == '未提出' else ''] * len(row)
                                
                                st.dataframe(df_status.style.apply(highlight_unsubmitted, axis=1), hide_index=True, use_container_width=True)
                            else:
                                st.error('スプレッドシートの「名簿」タブに「従業員コード」「名前」「部門」の列が見つかりません。')
                        else:
                            st.warning('スプレッドシートの「名簿」タブにスタッフのデータが登録されていません。（2行目以降が空です）')
                    else:
                        st.error('データの取得に失敗しました。')
                except Exception as e:
                    st.error(f'読み込みエラー: {e}')

# ==========================================
# メインの画面描画
# ==========================================
init_session_state()
input_disabled = st.session_state.confirm_mode or st.session_state.is_submitted
num_days, day_labels = get_month_days()

# あらかじめ全日程の変数を希望なしとして作成（一括入力のため）
for label in day_labels:
    if f'radio_{label}' not in st.session_state:
        st.session_state[f'radio_{label}'] = '希望なし'
    if f'time_{label}' not in st.session_state:
        st.session_state[f'time_{label}'] = ''

# --- タイトル 管理者パネル ---
col_title, col_admin = st.columns([4, 1])
with col_title:
    st.markdown(f'### {TARGET_YEAR}年{TARGET_MONTH}月分 シフト希望提出フォーム')
    st.write('時間の希望がある日だけ選択してください。希望がない日は「希望なし」のままでOKです。\n\nシステムの不具合がありましたら、不具合の画面の写真とともに、箭内にご連絡ください。')
with col_admin:
    show_admin_panel()

st.divider()

# メンテナンス期間中は表示しないようにする
if 0 <= now_jst.hour < 4:
    st.error('**現在システムメンテナンス中です**\n\n毎日 **00:00 〜 04:00** はメンテナンスのためご利用いただけません。\n恐れ入りますが、この時間を避けて再度アクセスしてください。')
    st.stop()

# --- 入力エリア ---
col_left, col_right = st.columns([1, 2])

with col_left:
    st.subheader('基本情報')
    name = st.text_input('1. お名前（フルネーム）', key='input_name', disabled=input_disabled)
    emp_code = st.text_input('2. 従業員コード（数字）', key='input_code', disabled=input_disabled)
    department = st.selectbox('3. 部門を選択', DEPARTMENTS, key='input_dept', disabled=input_disabled)
    target_days = st.number_input('4. 希望出勤時間（希望がない場合は0を入力してください）', min_value=0, max_value=120, value=0, step=1, key='input_days', disabled=input_disabled)
    remarks = st.text_area('5. 備考（自由記述）', key='input_remarks', disabled=input_disabled, placeholder='テスト期間や、時間指定の補足などがあれば記入してください。')

with col_right:
    # 日ごとの希望と一括入力の2つの欄を作成
    col_sub, col_btn = st.columns([2, 1])
    with col_sub:
        st.subheader('日ごとの希望')
    with col_btn:
        if st.button('一括入力する', use_container_width=True, disabled=input_disabled):
            bulk_input_dialog(day_labels)

    tab_titles = ['1日〜7日', '8日〜14日', '15日〜21日', '22日〜28日']
    day_groups = [day_labels[0:7], day_labels[7:14], day_labels[14:21], day_labels[21:28]]
    if num_days > 28:
        tab_titles.append(f'29日〜{num_days}日')
        day_groups.append(day_labels[28:num_days])
    
    tabs = st.tabs(tab_titles)
    shift_requests = {}
    
    for i, tab in enumerate(tabs):
        with tab:
            for label in day_groups[i]:
                choice = st.radio(f'**{label}**', ['希望なし', '休', '早', '遅', '時間指定', '有給'], horizontal=True, key=f'radio_{label}', disabled=input_disabled)
                if choice == '時間指定':
                    specific_time = st.text_input(f'↳ 【{label}】希望時間を入力（例：11-19, 早-15, 14-L）', key=f'time_{label}', disabled=input_disabled)
                    shift_requests[label] = specific_time if specific_time else '時間指定(未入力)'
                elif choice == '希望なし':
                    shift_requests[label] = '出' 
                elif choice == '有給':
                    shift_requests[label] = '有'
                else:
                    shift_requests[label] = choice
                st.write('---')

st.divider()

# --- プレビュー ボタンエリア ---
if st.session_state.is_submitted:
    st.success(f'{name}さん、シフトの提出が完了しました。')
    st.info('※修正が必要な場合は店長または箭内へ連絡してください。')
    st.markdown('#### 提出されたシフト内容')
    st.table(generate_styled_calendar(day_labels, shift_requests))

elif st.session_state.confirm_mode:
    st.warning('以下の内容で確定してよろしいですか？')
    st.markdown('#### シフト希望プレビュー')
    st.table(generate_styled_calendar(day_labels, shift_requests))

    if remarks:
        st.markdown('**【備考】**')
        st.write(remarks)
    
    col_btn1, col_btn2 = st.columns(2)
    with col_btn1:
        if st.button('戻って修正する', use_container_width=True):
            st.session_state.confirm_mode = False
            st.rerun()
    with col_btn2:
        if st.button('この内容で確定・提出する', type='primary', use_container_width=True):
            # 連打防止
            st.session_state.is_processing = True
            st.rerun()

# 連打対策
if st.session_state.get('is_processing', False) and not st.session_state.is_submitted:
    with st.spinner('クラウドにシフトを送信しています...'):
        result = save_shift_data(emp_code, name, department, target_days, shift_requests, remarks)
    
    if result == 'gas_error':
        st.error('【通信エラー】Googleスプレッドシートへの送信に失敗しました。時間をおいてもう一度お試しいただくか、管理者へ連絡してください。')
        st.session_state.is_processing = False
    else:
        st.session_state.is_submitted = True
        st.session_state.confirm_mode = False
        st.session_state.is_processing = False
        st.rerun()

else:
    if not st.session_state.is_submitted and not st.session_state.confirm_mode:
        st.markdown('#### ライブプレビュー')
        st.table(generate_styled_calendar(day_labels, shift_requests))
        st.write('') 
        if st.button('確認画面へ進む', type='primary', use_container_width=True):
            if not name: st.error('お名前が入力されていません。')
            if department == '選択してください': st.error('部門が選択されていません。')
            if not emp_code: st.error('従業員コードが入力されていません。')
            
            if name and department != '選択してください' and emp_code:
                st.session_state.confirm_mode = True
                st.rerun()
